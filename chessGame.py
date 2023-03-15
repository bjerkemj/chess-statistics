#Tinus Alsos og Johan Bjerkem
import os
import re
import xlsxwriter
import openpyxl
ROOT = os.path.dirname(os.path.abspath(__file__))

class ChessGame:
    requiredTags = ["Event", "Site", "Date",
                    "Round", "White", "Black", "Result"]
    optionalTags = ["ECO", "Opening", "Variation", "PlyCount", "WhiteElo", "BlackElo"]

    def __init__(self, pgn: str = None, xlsxName: str = None) ->  None:
        self.metaData = {}
        self.moves = []
        self.stockfishWhite = False
        self.draw = False
        self.whiteWon = False
        self.totalMoves = 0
        if pgn != None:
            self.loadPgn(pgn)
        else:
            self.loadXlsx(xlsxName)

    def getStockfishWhite(self) -> bool:
        return self.stockfishWhite

    def getDraw(self) -> bool:
        return self.draw

    def getOpening(self) -> str:
        return self.metaData['Opening']

    def getWhiteWon(self) -> bool:
        return self.whiteWon

    def getTotalMoves(self) -> int:
        return self.totalMoves

    def getMoveByNumber(self, moveNumber: int) -> str:
        if type(moveNumber) == str:
            moveNumber = int(moveNumber)
        return self.moves[moveNumber - 1][0] if moveNumber <= self.totalMoves else False

    def getStockfishWon(self) -> bool:
        return self.getStockfishWhite() == self.getWhiteWon() and not self.getDraw()

    def getStockfishLost(self) -> bool:
        return self.getStockfishWhite() != self.getWhiteWon() and not self.getDraw()

    def getStockfishWhiteWin(self) -> bool:
        return self.getStockfishWhite() and self.getStockfishWon()

    def getStockfishWhiteLoss(self) -> bool:
        return self.getStockfishWhite() and self.getStockfishLost()

    def getStockfishWhiteDraw(self) -> bool:
        return self.getStockfishWhite() and self.getDraw()

    def getStockfishBlackWin(self) -> bool:
        return not self.getStockfishWhite() and self.getStockfishWon()

    def getStockfishBlackLoss(self) -> bool:
        return not self.getStockfishWhite() and self.getStockfishLost()

    def getStockfishBlackDraw(self) -> bool:
        return not self.getStockfishWhite() and self.getDraw()

    def loadPgn(self, pgn: str) -> None:
        split = pgn.split("\n")
        dataSplitIndex = split.index("")
        metaDataText = " ".join(split[:dataSplitIndex])
        gameDataText = " ".join(split[dataSplitIndex:]).strip()

        for tag in self.requiredTags:
            tagIndex = metaDataText.find(tag)
            firstQuoteIndex = pgn.find("\"", tagIndex)
            secondQuoteIndex = pgn.find("\"", firstQuoteIndex+1)
            self.metaData[tag] = pgn[firstQuoteIndex+1:secondQuoteIndex]

        for tag in self.optionalTags:
            if tag in metaDataText:
                tagIndex = metaDataText.find(tag)
                firstQuoteIndex = pgn.find("\"", tagIndex)
                secondQuoteIndex = pgn.find("\"", firstQuoteIndex+1)
                self.metaData[tag] = pgn[firstQuoteIndex+1:secondQuoteIndex]

        while "{" in gameDataText:
            moveStartIndex = gameDataText.find(" ") + 1
            moveEndIndex = gameDataText.find("{") - 1
            moveInfo = gameDataText[moveStartIndex:moveEndIndex]

            commentStartIndex = gameDataText.find("{") + 1
            commentEndIndex = gameDataText.find("}")
            commentInfo = gameDataText[commentStartIndex:commentEndIndex]

            self.moves.append([moveInfo, commentInfo])

            gameDataText = gameDataText[commentEndIndex +
                                        (1 if len(self.moves) % 2 == 1 else 2):]

        self.updateGameData()

    def savePgn(self, saveName: str, type: str = "w") -> None:
        with open(os.path.join(ROOT, saveName + ".pgn"), type + '+') as f:
            for key, value in self.metaData.items():
                f.write(f"[{key} \"{value}\"]\n")
            f.write("\n")

            gameDataText = ""
            for i, moveInfo in enumerate(self.moves):
                if i % 2 == 0:
                    gameDataText += str(int(i/2+1)) + ". "
                gameDataText += f"{moveInfo[0]} "
                gameDataText += "{"
                gameDataText += moveInfo[1]
                gameDataText += "} "

            index = -1
            gameDataText += self.metaData["Result"] + "\n"
            lineLength = 81
            while index+lineLength < len(gameDataText):
                if index == -1:
                    spaceIndex = gameDataText.rfind(" ", 0, index+lineLength)
                else:
                    spaceIndex = gameDataText.rfind(" ", index, index+lineLength)
                gameDataList = list(gameDataText)
                gameDataList[spaceIndex] = "\n"
                gameDataText = "".join(gameDataList)
                index = spaceIndex 
            gameDataText += "\n"            
            f.writelines(gameDataText)

    def saveXlsx(self, saveName: str) -> None:
        os.system('rm ' + saveName + '.xlsx')
        workbook = xlsxwriter.Workbook(saveName + '.xlsx')
        worksheet = workbook.add_worksheet()

        row = 0
        for key, value in self.metaData.items():
                worksheet.write(row, 0, key)
                worksheet.write(row, 1, value)
                row += 1

        row += 1
        column = 0
        for move in self.moves:
            worksheet.write(row, column, move[0])
            worksheet.write(row, column+1, move[1])
            if column == 0:
                column = 2
            else:
                column = 0
                row+=1

        workbook.close()

    def loadXlsx(self, xlsxName: str):
        wookbook = openpyxl.load_workbook(xlsxName + '.xlsx', data_only=True)
        worksheet = wookbook.active
        pgnString = ""

        for i, row in enumerate(worksheet.iter_rows()):
            if row[0].internal_value == None:
                datasplit = i
                break
            pgnString += "[" + row[0].internal_value + ' "' + row[1].internal_value +'"' + "]\n"

        pgnString += "\n"

        i = 1
        for row in worksheet.iter_rows(min_row=datasplit+2):
            pgnString += str(i) + ". " + row[0].internal_value + " {" + row[1].internal_value + "} "
            if row[2].internal_value != None:
                pgnString += row[2].internal_value + " {" + row[3].internal_value + "} "
            i+=1
        
        self.loadPgn(pgnString)

    def updateGameData(self) -> None:
        self.stockfishWhite = True if re.search(
            r"\bStockfish\b", self.metaData["White"]) else False
        self.totalMoves = int(self.metaData["PlyCount"])
        if self.metaData["Result"] == "1/2-1/2":
            self.draw = True
            return
        self.whiteWon = True if self.metaData["Result"] == "1-0" else False
